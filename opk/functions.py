#!/usr/bin/env python3
# coding: utf-8
#pylint: disable=invalid-name,import-error
"""Common functions for OPK"""

from openpyxl import load_workbook
from opk.config import COLUMNS, USED_COLUMNS, MEMBER_TYPES

def first_name(data):
    """Fetch all but first work from string"""
    return ' '.join(data.split()[1:])

def last_name(data):
    """Fetch first word from string"""
    return ''.join(data.split()[:1])

def get_cell_value(ws, field_name, row):
    """Take a string and extract its cell value from row

    Takes the ws, column name, eg. member_id and row number
    as args.
    """
    field_value = ws["{}{}".format(COLUMNS[field_name], row)].value
    # Kludge because names are inappropriately formatted as "LastName Name Name Name"
    # Which isn't very pretty.
    if field_name == 'name':
        return "{} {}".format(first_name(field_value), last_name(field_value))
    else:
        return field_value


def gen_memberlist(filename, member_types=MEMBER_TYPES):
    """Returns list of members"""
    members = list()
    wb = load_workbook(filename=filename, read_only=False)
    ws = wb.active
    for row in range(2, ws.max_row+1):
        member = dict()
        for field in USED_COLUMNS:
            member[field] = get_cell_value(ws, field, row)
        if member['member_type'] in member_types:
            members.append(member)
    return members
